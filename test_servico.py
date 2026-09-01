import os
import shutil
import tempfile
import unittest
from datetime import time as dtime
from unittest.mock import patch

from dotenv import load_dotenv
from openpyxl import load_workbook

import create_excel
import main
from api_request import ApiRequest


class EnvMixin:
    def setUp(self):
        self._old_env = os.environ.copy()

    def tearDown(self):
        os.environ.clear()
        os.environ.update(self._old_env)


class TestFlagsEAgenda(EnvMixin, unittest.TestCase):
    def test_is_enabled_padrao_e_aliases(self):
        os.environ.pop("CADASTRO_ENABLED", None)
        self.assertTrue(main._is_enabled("CADASTRO_ENABLED", True))
        self.assertFalse(main._is_enabled("CADASTRO_ENABLED", False))

        os.environ["CADASTRO_ENABLED"] = "true"
        self.assertTrue(main._is_enabled("CADASTRO_ENABLED", False))
        os.environ["CADASTRO_ENABLED"] = "SIM"
        self.assertTrue(main._is_enabled("CADASTRO_ENABLED", False))
        os.environ["CADASTRO_ENABLED"] = "false"
        self.assertFalse(main._is_enabled("CADASTRO_ENABLED", True))
        os.environ["CADASTRO_ENABLED"] = "nao"
        self.assertFalse(main._is_enabled("CADASTRO_ENABLED", True))

        os.environ["CADASTRO_ENABLED"] = "talvez"
        with self.assertRaises(ValueError):
            main._is_enabled("CADASTRO_ENABLED")

    def test_parse_dias_e_horarios(self):
        self.assertEqual(main._parse_schedule_days("SEG", "CADASTRO_SCHEDULE_DAYS"), [0])
        self.assertEqual(
            main._parse_schedule_days("seg, TER,seg", "SCHEDULE_DAYS"), [0, 1]
        )
        with self.assertRaises(ValueError):
            main._parse_schedule_days("", "SCHEDULE_DAYS")
        with self.assertRaises(ValueError):
            main._parse_schedule_days("SEGUNDA", "SCHEDULE_DAYS")

        times = main._parse_schedule_times("12:00,08:00,12:00", "SCHEDULE_TIMES")
        self.assertEqual(times, [dtime(8, 0), dtime(12, 0)])
        with self.assertRaises(ValueError):
            main._parse_schedule_times("8h", "CADASTRO_SCHEDULE_TIMES")

    def test_timezone_obrigatorio(self):
        os.environ["TIMEZONE"] = ""
        with self.assertRaises(ValueError):
            main._get_timezone()
        os.environ["TIMEZONE"] = "America/Sao_Paulo"
        tz, name = main._get_timezone()
        self.assertEqual(name, "America/Sao_Paulo")
        self.assertTrue(tz)

    def test_nome_ga(self):
        self.assertEqual(main._nome_ga("12 - João Silva"), "JOÃO SILVA")
        self.assertEqual(main._nome_ga("SEM SEPARADOR"), "SEM SEPARADOR")

    def test_build_jobs_independentes(self):
        os.environ["INADIMPLENCIA_ENABLED"] = "true"
        os.environ["CADASTRO_ENABLED"] = "true"
        os.environ["SCHEDULE_DAYS"] = "SEG"
        os.environ["SCHEDULE_TIMES"] = "12:00"
        os.environ["CADASTRO_SCHEDULE_DAYS"] = "SEG"
        os.environ["CADASTRO_SCHEDULE_TIMES"] = "08:00"

        jobs = main._build_jobs()
        self.assertEqual([j["name"] for j in jobs], ["inadimplencia", "cadastro_incompleto"])
        self.assertEqual(jobs[0]["times"][0], dtime(12, 0))
        self.assertEqual(jobs[1]["times"][0], dtime(8, 0))

        os.environ["INADIMPLENCIA_ENABLED"] = "false"
        jobs = main._build_jobs()
        self.assertEqual([j["name"] for j in jobs], ["cadastro_incompleto"])

        os.environ["CADASTRO_ENABLED"] = "false"
        with self.assertRaises(ValueError):
            main._build_jobs()

    def test_cadastro_usa_default_segunda_08(self):
        os.environ["INADIMPLENCIA_ENABLED"] = "false"
        os.environ["CADASTRO_ENABLED"] = "true"
        os.environ.pop("CADASTRO_SCHEDULE_DAYS", None)
        os.environ.pop("CADASTRO_SCHEDULE_TIMES", None)
        jobs = main._build_jobs()
        self.assertEqual(jobs[0]["name"], "cadastro_incompleto")
        self.assertEqual(jobs[0]["days"], [0])
        self.assertEqual(jobs[0]["times"][0], dtime(8, 0))

    def test_run_mode_invalido(self):
        os.environ["TIMEZONE"] = "America/Sao_Paulo"
        os.environ["RUN_MODE"] = "cron"
        with self.assertRaises(ValueError):
            main.main()

    def test_job_falha_nao_impede_proximo(self):
        executed = []
        main._executar_job(
            {"name": "fail", "run": lambda: (_ for _ in ()).throw(RuntimeError("boom"))}
        )
        main._executar_job({"name": "ok", "run": lambda: executed.append("ok")})
        self.assertEqual(executed, ["ok"])


class TestExcel(unittest.TestCase):
    def setUp(self):
        self.generated = []

    def tearDown(self):
        for path in self.generated:
            if os.path.exists(path):
                try:
                    os.remove(path)
                except OSError:
                    pass

    def _track(self, path):
        self.generated.append(path)
        return path

    def test_listagem_sem_colunas_financeiras(self):
        path = self._track(
            create_excel.writeExcel(
                [
                    {
                        "codcli": 1,
                        "cliente": "A",
                        "codrca": 10,
                        "historico": "CADASTRO INCOMPLETO",
                    },
                    {
                        "codcli": 2,
                        "cliente": "B",
                        "codrca": 20,
                        "historico": "CADASTRO INCOMPLETO",
                    },
                ],
                "teste-unit-listagem",
                modo="listagem",
            )
        )
        self.assertTrue(os.path.exists(path))
        wb = load_workbook(path)
        try:
            values = [cell.value for row in wb.active.iter_rows(min_col=1, max_col=1) for cell in row]
            self.assertIn("CODCLI", values)
            self.assertTrue(any(isinstance(v, str) and v.startswith("RCA:") for v in values))
        finally:
            wb.close()

    def test_listagem_sem_codrca(self):
        path = self._track(
            create_excel.writeExcel(
                [{"codcli": 9, "cliente": "Z"}],
                "teste-unit-sem-rca",
                modo="listagem",
            )
        )
        wb = load_workbook(path)
        try:
            self.assertEqual(wb.active["A1"].value, "CODCLI")
            self.assertEqual(wb.active["B2"].value, "Z")
        finally:
            wb.close()

    def test_financeiro_ainda_soma(self):
        path = self._track(
            create_excel.writeExcel(
                [
                    {
                        "codsupervisor": 1,
                        "codrca": 10,
                        "valor_total_com_juros": "10.5",
                        "valor_total_original": "10",
                    }
                ],
                "teste-unit-financeiro",
                modo="financeiro",
            )
        )
        self.assertTrue(os.path.exists(path))

    def test_modo_e_extensao_invalidos(self):
        with self.assertRaises(ValueError):
            create_excel.writeExcel([{"a": 1}], "x", modo="pdf")
        with self.assertRaises(ValueError):
            create_excel.writeExcel([{"a": 1}], "x", extension="pdf", modo="listagem")


class TestLimpezaArquivos(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self._cwd = os.getcwd()
        os.chdir(self.tmpdir)
        os.makedirs("arquivos-gerados", exist_ok=True)
        open(os.path.join("arquivos-gerados", "sup-keep.xlsx"), "w").close()
        open(os.path.join("arquivos-gerados", "cadastro-sup-tmp.xlsx"), "w").close()

    def tearDown(self):
        os.chdir(self._cwd)
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_limpa_so_prefixo_do_job(self):
        main._limpar_arquivos_gerados("cadastro-sup-")
        self.assertTrue(os.path.exists(os.path.join("arquivos-gerados", "sup-keep.xlsx")))
        self.assertFalse(
            os.path.exists(os.path.join("arquivos-gerados", "cadastro-sup-tmp.xlsx"))
        )


class TestJwt(unittest.TestCase):
    def test_token_tem_tres_partes(self):
        load_dotenv()
        api = ApiRequest()
        token = api.generate_token_request(
            {"database": "atacado", "matricula": "3312", "indexPage": 0}
        )
        self.assertEqual(token.count("."), 2)
        self.assertTrue(len(token) > 20)


class TestDryRunNaoEnviaWhatsapp(EnvMixin, unittest.TestCase):
    def test_dry_run_gera_excel_e_nao_chama_wpp(self):
        os.environ["DRY_RUN"] = "true"
        os.environ["TIMEZONE"] = "America/Sao_Paulo"
        supervisores = [
            {"codigo": "99", "titulo": "99 - TESTE GA", "telefone": "33999999999"}
        ]
        relatorio = [{"codcli": 1, "cliente": "Cliente Teste", "codrca": 10}]

        with patch.object(ApiRequest, "get_supervisores_ativos", return_value=supervisores), patch.object(
            ApiRequest,
            "relatorio_cadastro_incompleto_filtrando_supervisor",
            return_value=relatorio,
        ), patch.object(
            ApiRequest, "send_mensagem_chatbot", return_value="NAO DEVERIA CHAMAR"
        ) as send_mock:
            main.run_job_cadastro_incompleto()
            send_mock.assert_not_called()

        esperado = os.path.join(
            "arquivos-gerados", f"cadastro-sup-99-{__import__('datetime').datetime.now().date()}.xlsx"
        )
        self.assertTrue(os.path.exists(esperado))
        os.remove(esperado)


class TestLiveApi(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        load_dotenv()
        cls.api = ApiRequest()
        cls.supervisores = cls.api.get_supervisores_ativos()

    def test_supervisores_ativos(self):
        self.assertTrue(self.supervisores, "API de supervisores retornou vazio")
        self.assertIsInstance(self.supervisores, list)
        primeiro = self.supervisores[0]
        for campo in ("codigo", "titulo"):
            self.assertIn(campo, primeiro)

    def test_cadastro_incompleto_contrato(self):
        alvo = next((s for s in self.supervisores if s.get("codigo") is not None), None)
        self.assertIsNotNone(alvo)
        dados = self.api.relatorio_cadastro_incompleto_filtrando_supervisor(
            alvo["codigo"], alvo["titulo"]
        )
        self.assertIsInstance(dados, list)
        if dados:
            self.assertIsInstance(dados[0], dict)
            path = create_excel.writeExcel(
                dados,
                f"teste-live-cadastro-{alvo['codigo']}",
                modo="listagem",
            )
            self.assertTrue(os.path.exists(path))
            wb = load_workbook(path)
            try:
                self.assertGreaterEqual(wb.active.max_row, 2)
            finally:
                wb.close()
            os.remove(path)

    def test_inadimplencia_ainda_responde(self):
        alvo = next((s for s in self.supervisores if s.get("codigo") is not None), None)
        self.assertIsNotNone(alvo)
        dados = self.api.relatorio_inadiplencia_filtrando_supervisor(
            alvo["codigo"], alvo["titulo"]
        )
        self.assertIsInstance(dados, list)

    def test_whatsapp_api_alcancavel(self):
        import requests

        url = os.getenv("WPP_API_URL", "").rstrip("/")
        self.assertTrue(url, "WPP_API_URL não definido")
        try:
            response = requests.get(url, timeout=8)
        except requests.RequestException as error:
            self.fail(f"WhatsApp API inacessível em {url}: {error}")
        self.assertLess(response.status_code, 500, f"WPP status {response.status_code}")


if __name__ == "__main__":
    unittest.main(verbosity=2)
