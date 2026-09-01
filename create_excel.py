import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os


def _ajustar_largura_colunas(ws):
     for col in ws.columns:
          max_len = 0
          col_letter = col[0].column_letter
          for cell in col:
               if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
          ws.column_dimensions[col_letter].width = max_len + 2


def _coluna_existente(df: pd.DataFrame, nomes: tuple[str, ...]):
     for nome in nomes:
          if nome in df.columns:
               return nome
     return None


def _adicionar_blocos_por_rca(ws, df: pd.DataFrame, linha_inicio: int) -> int:
     coluna_rca = _coluna_existente(df, ('CODRCA', 'COD_RCA'))
     if not coluna_rca:
          return linha_inicio

     grouped = df.groupby(coluna_rca, dropna=False)
     for rca_name, group_df in grouped:
          linha_inicio += 1
          linha_inicio = criar_tabela_agrupada(
               ws,
               group_df.reset_index(drop=True),
               linha_inicio,
               titulo=f"RCA: {rca_name}",
               colunas_soma=[],
          )
     return linha_inicio


def writeExcel(data: list, file_name: str, extension: str = 'xlsx', modo: str = 'financeiro'):
     os.makedirs('arquivos-gerados', exist_ok=True)

     df = pd.DataFrame(data)

     df.columns = df.columns.str.upper()

     if modo not in ('financeiro', 'listagem'):
          raise ValueError("Modo não suportado, use financeiro ou listagem")

     if extension == 'csv':
          out_path = f'arquivos-gerados/{file_name}.{extension}'
          df.to_csv(out_path, index=False)
          return out_path

     if extension != 'xlsx':
          raise ValueError('Extensão não suportada, use xlsx ou csv')

     caminho_arquivo = f'arquivos-gerados/{file_name}.{extension}'
     df.to_excel(caminho_arquivo, index=False)
     wb = load_workbook(caminho_arquivo)
     try:
          ws = wb['Sheet1']

          _ajustar_largura_colunas(ws)

          if modo == 'listagem':
               fonte_cabecalho = Font(bold=True)
               for cell in ws[1]:
                    cell.font = fonte_cabecalho
                    cell.alignment = Alignment(horizontal="center")

               wb.save(caminho_arquivo)
               return caminho_arquivo

          colunas_para_soma = ['VALOR_TOTAL_COM_JUROS', 'VALOR_TOTAL_ORIGINAL']

          for colunas in colunas_para_soma:
               df[colunas] = pd.to_numeric(df[colunas], errors='coerce')

          ws["A1"].font = Font(bold=True)
          ws["A1"].alignment = Alignment(horizontal="center")

          ultima_linha = ws.max_row + 2

          coluna_supervisor = _coluna_existente(df, ('CODSUPERVISOR', 'COD_SUPERVISOR'))
          if coluna_supervisor:
               valores_agrupados = df.groupby(coluna_supervisor, dropna=False)[colunas_para_soma].sum().reset_index()
               ultima_linha = criar_tabela_agrupada(
                    ws,
                    valores_agrupados,
                    ultima_linha,
                    titulo=f"TOTAL GERAL POR '{coluna_supervisor}'",
                    colunas_soma=colunas_para_soma,
               )

          coluna_rca = _coluna_existente(df, ('CODRCA', 'COD_RCA'))
          if coluna_rca:
               tabela_rca = df.groupby(coluna_rca, dropna=False)[colunas_para_soma].sum().reset_index()
               tabela_rca = tabela_rca[[coluna_rca, *colunas_para_soma]]
               ultima_linha += 1
               ultima_linha = criar_tabela_agrupada(
                    ws,
                    tabela_rca,
                    ultima_linha,
                    titulo=f"TOTAL POR {coluna_rca}",
                    colunas_soma=colunas_para_soma,
               )

          _adicionar_blocos_por_rca(ws, df, ultima_linha)

          wb.save(caminho_arquivo)
          return caminho_arquivo
     finally:
          wb.close()
     

def criar_tabela_agrupada(planilha, dados: pd.DataFrame, linha_inicio: int, titulo: str, colunas_soma: list[str]):
     
     # Configurações de estilização de quadro totalizador
     espessura = Side(border_style="thin", color="000000")
     borda = Border(left=espessura, right=espessura, top=espessura, bottom=espessura)
     cor_cabecalho = PatternFill("solid", fgColor="DDDDDD")
     fonte_negrito = Font(bold=True)
     cor_titulo = PatternFill("solid", fgColor="0CCC82")

     total_cols = len(dados.columns)
     if titulo:
          planilha.merge_cells(start_row=linha_inicio, start_column=1, end_row=linha_inicio, end_column=total_cols)
          tcell = planilha.cell(row=linha_inicio, column=1, value=titulo)
          tcell.font = fonte_negrito
          tcell.fill = cor_titulo
          tcell.alignment = Alignment(horizontal="left")
          tcell.border = borda
          linha_inicio += 1

     for index, coluna in enumerate(dados.columns, start=1):
          hcell = planilha.cell(row=linha_inicio, column=index, value=coluna)
          hcell.font = fonte_negrito
          hcell.fill = cor_cabecalho
          hcell.alignment = Alignment(horizontal="center")
          hcell.border = borda

     for i, linha in dados.iterrows():
          for j, coluna in enumerate(dados.columns, start=1):
               
               if type(i) != int:
                    return linha_inicio
               
               proxima_linha = linha_inicio + 1 + i

               if coluna in colunas_soma:
                    cell = planilha.cell(row=proxima_linha, column=j, value=pd.to_numeric(linha[coluna]))
               
               cell = planilha.cell(row=proxima_linha, column=j, value=linha[coluna])
               
               cell.border = borda


     return linha_inicio + 1 + len(dados)
